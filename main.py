import asyncio
import email
import imaplib
import logging
import os
import zipfile
from email.header import decode_header
import openpyxl
import pandas as pd
from multiprocessing import Process


# Логирование
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("app.log", encoding="utf-8"),  # Логи сохраняются в файл
        logging.StreamHandler()  # Логи выводятся на консоль
    ]
)

def get_untouchable_articles(file_path) -> list[str]:
    """Возвращает список неприкасаемых артикулов."""
    untouched_articles = []

    with open(file_path, 'r', encoding='utf-8') as file:
        articles_text = file.readline()
        for article in articles_text.split(', '):
            untouched_articles.append(article)

    return untouched_articles


class MailHandler:
    """Подключается к почте"""

    def __init__(self, email_address, password):
        self.imap = imaplib.IMAP4_SSL('imap.yandex.ru')
        self.imap.login(email_address, password)

    def download_and_get_file_name(self, folder='INBOX', download_folder='attachments') -> list[str]:
        """Берет последнее письмо полученное, скачивает файлы и возвращает их имена"""
        # Переход в папку
        print("[+]Ищем письмо на почте[+]")
        logging.info("[+]Ищем письмо на почте[+]")
        self.imap.select(folder)

        # Поиск всех писем
        status, messages = self.imap.search(None, 'ALL')
        if status != "OK":
            logging.error("Письмо не найдено")
            print("Ошибка поиска писем.")
            return []

        # Берем ID последнего сообщения
        messages = messages[0].split()
        if not messages:
            logging.error("Нет писем в папке.")
            print("Нет писем в папке.")
            return []

        latest_msg_id = messages[-1]

        # Получаем письмо
        res, msg = self.imap.fetch(latest_msg_id, '(RFC822)')
        if res != 'OK':
            print(f"Ошибка получения письма {latest_msg_id}")
            return []

        # Разбираем письмо
        msg = email.message_from_bytes(msg[0][1])
        subject, encoding = decode_header(msg["Subject"])[0]
        if isinstance(subject, bytes):
            subject = subject.decode(encoding if encoding else 'utf-8')
        logging.info(f"Обрабатываем письмо с темой: {subject}")
        print(f"Обрабатываем письмо с темой: {subject}")

        # Создаем папку для загрузки, если её нет
        if not os.path.exists(download_folder):
            os.makedirs(download_folder)

        excel_files = []  # Список для хранения Excel файлов

        # Обрабатываем вложения
        for part in msg.walk():
            if part.get_content_disposition() == 'attachment':
                filename = part.get_filename()
                if filename:
                    filename = decode_header(filename)[0][0]
                    if isinstance(filename, bytes):
                        filename = filename.decode()

                    filepath = os.path.join(download_folder, filename)
                    with open(filepath, "wb") as f:
                        f.write(part.get_payload(decode=True))

                    # Проверяем, является ли файл архивом .zip
                    if zipfile.is_zipfile(filepath):
                        with zipfile.ZipFile(filepath, 'r') as zip_ref:
                            zip_ref.extractall(download_folder)
                            print(f"Архив {filepath} распакован.")

                        # Получаем список файлов из архива и проверяем на наличие Excel файлов
                        extracted_files = zip_ref.namelist()
                        for file_name in extracted_files:
                            if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                                excel_files.append(file_name)

                        # Удаляем архив после распаковки
                        os.remove(filepath)
                    else:
                        print(f"Файл не является архивом: {filename}")

        return excel_files  # Возвращаем все найденные Excel файлы

    def close_connection(self):
        self.imap.logout()


class ExcelHandler:
    """Обработчик файла с прайсами склада."""

    @classmethod
    def convert_to_dataset(cls, file_path) -> dict:
        """Открывает, читает, а затем приводит к удобочитаемому формату,
         где ключом будет артикул поставщика(article), а значением словарь с его данными."""

        try:
            dataframe = pd.read_excel(
                file_path,
                header=None,
                names=['brand', 'article_dont_used', "name", "article", "seller_article", "stock", "price"],
                index_col=None,
                engine='openpyxl'
            )
            result = {}

            for _, row in dataframe.iterrows():
                result[row["seller_article"]] = {
                    'brand': row["brand"],
                    "name": row["name"],
                    "article": row["article"],
                    "stock": row["stock"],
                    "price": row["price"]
                }

            print('[+]Получили файл[+]')
            return result
        except FileNotFoundError:
            raise FileNotFoundError(f'File: "{file_path}" not found')


    @classmethod
    def get_final_report(cls, first_sklad_dataset: dict, second_sklad_dataset: dict) -> dict:
        """Получение результирующего датасета."""
        result = {}

        # Преобразуем артикулы в множества для быстрого поиска
        first_sklad_articles = set(first_sklad_dataset.keys())
        second_sklad_articles = set(second_sklad_dataset.keys())

        # Вспомогательная функция для расчета цены
        def calculate_price(price: int) -> int:
            return price * 3 if price < 300 else price * 2

        # Обрабатываем все артикулы
        all_articles = first_sklad_articles.union(second_sklad_articles)

        for article in all_articles:
            first_item = first_sklad_dataset.get(article, {})
            second_item = second_sklad_dataset.get(article, {})

            # Объединяем данные двух складов
            name = first_item.get("name", second_item.get("name", "Unknown"))
            brand = first_item.get("brand", second_item.get("brand", "Unknown"))
            stock1 = first_item.get("stock", "0")
            stock2 = second_item.get("stock", "0")
            price1 = int(first_item.get("price", 0))
            price2 = int(second_item.get("price", 0))

            # Рассчитываем stock
            if '>' in stock1 or '>' in stock2:
                stock = stock1 if '>' in stock1 else stock2
            else:
                stock = int(stock1) + int(stock2)

            # Рассчитываем цену
            max_price = max(price1, price2)
            final_price = calculate_price(max_price)

            # Добавляем данные в результат
            result[article] = {
                "name": name,
                "article": article,
                "stock": stock,
                "brand": brand,
                "price": final_price
            }

        print('[+]Получен отчет слияния[+]')
        return result

    @classmethod
    def write_to_stocks_pattern(cls, report: dict, untouchable_articles: list[str]) -> None:
        wb = openpyxl.load_workbook('stocks_pattern.xlsx')
        sheet = wb.active


        for row in sheet.iter_rows(max_row=2, max_col=sheet.max_column, min_row=sheet.max_row):
            article = row[6]
            if article in untouchable_articles:
                continue
            else:
                row[1].value = 0

        # Читаем все строки Excel и создаем словарь с артикулами как ключами
        article_to_row = {}
        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
            article_cell = row[6]
            if article_cell.value:
                article_to_row[article_cell.value] = row

        # Записываем цену для каждого артикула
        for article, data in report.items():
            if article in article_to_row:
                row = article_to_row[article]
                row[1].value = data['stock']  # 2-я колонка (index 1)
                logging.info(f"[+]Обновлены остатки: артикул {article} в строке {row[1].row}[+]")
                print(f"[+]Обновлены остатки: артикул {article} в строке {row[1].row}[+]")
            else:
                logging.info(f"Артикул {article} не найден, пропускаем.")
            #     print(f"Артикул {article} не найден, пропускаем.")

        # Сохраняем изменения в новый файл
        wb.save('stocks.xlsx')
        print("[+]Файл stocks.xlsx обновлен.[+]")

    @classmethod
    def write_to_price_pattern(cls, report: dict) -> None:
        wb = openpyxl.load_workbook('price_pattern.xlsx')
        sheet = wb.active

        # Читаем все строки Excel и создаем словарь с артикулами как ключами
        article_to_row = {}
        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
            article_cell = row[3]  # Артикул в колонке 4
            if article_cell.value:
                article_to_row[article_cell.value] = row

        # Записываем цену для каждого артикула
        for article, data in report.items():
            if article in article_to_row:
                row = article_to_row[article]
                row[9].value = data['price']  # 10-я колонка (index 9)
                logging.info(f"[+]Обновлена цена: артикул {article} в строке {row[3].row}[+]")
                print(f"[+]Обновлена цена: артикул {article} в строке {row[3].row}[+]")
            else:
                logging.info(f"[-]Артикул {article} не найден, пропускаем.[-]")
            #     print(f"[-]Артикул {article} не найден, пропускаем.[-]")

        # Сохраняем изменения в новый файл
        wb.save('price.xlsx')
        print("[+]Файл price.xlsx обновлен.[+]")


def main():
    yandex_email = 'avtowb@yandex.ru'
    yandex_app_password = 'fdkwtvzkcenjszje'

    y_mail = MailHandler(yandex_email, yandex_app_password)
    file_names = y_mail.download_and_get_file_name()

    untouchable_articles = get_untouchable_articles('untouchable_articles.txt')
    first_sklad_dataset = ExcelHandler.convert_to_dataset(file_path='attachments/' + file_names[0])
    second_sklad_dataset = ExcelHandler.convert_to_dataset(file_path='attachments/' + file_names[1])
    #
    report = ExcelHandler.get_final_report(first_sklad_dataset, second_sklad_dataset)

    process_write_to_stocks = Process(target=ExcelHandler.write_to_stocks_pattern, args=(report, untouchable_articles))
    process_write_to_price = Process(target=ExcelHandler.write_to_price_pattern, args=(report,))

    processes = [process_write_to_price, process_write_to_stocks]

    for process in processes:
        process.start()
        process.join()

    print('Программа успешно завершена')
    input()


if __name__ == '__main__':
    main()
